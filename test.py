from selenium import webdriver
import time
from secrets import secrets

opt = webdriver.ChromeOptions()
driver = webdriver.Chrome("webdriver/chromedriver", options=opt)

username = secrets['username']
pwd = secrets['password']

driver.get("https://www.surveymonkey.com/dashboard/")

print("Please log in now")

username_tf = driver.find_element_by_id('username')
pwd_tf = driver.find_element_by_id('password')
username_tf.send_keys(username)
pwd_tf.send_keys(pwd)

loginbtn = driver.find_elements_by_class_name('wds-button--arrow-right')[0]
loginbtn.click()

while not ("Welcome to SurveyMonkey!" in driver.title):
    time.sleep(0.5)
    pass

print("Logged in!")

driver.get("https://www.surveymonkey.com/analyze/browse/FgsgtgKNp6FH7sZFdIbZuJoG590A0xr9wKBkxY2YHF0_3D")

driver.implicitly_wait(50)

from openpyxl import Workbook
from bs4 import BeautifulSoup

wb = Workbook()
ws = wb.active

questions = []

for index in range(1, 36):

    # Go to respndant page
    respondBtn = driver.find_elements_by_class_name('respondent-goto-menu-btn')

    if len(respondBtn) == 0:
        print("Error! btn not found")
    
    respondBtn[0].click()
    gototf = driver.find_elements_by_class_name('goto-number-text')[0]
    gototf.clear()
    gototf.send_keys(index)
    gotobtn = driver.find_elements_by_class_name('goto-btn')[0]
    gotobtn.click()
    if index == 1:
        time.sleep(3)
    else:
        time.sleep(0.5)
    
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    divs = soup.findAll("div", {"class": "response-question"})[:10]
    print(divs)

    if not questions:
        questions = [div.find_all("div", {"class": "response-question-title-text"})[0] for div in divs]
        questions = [question.get_text().strip() for question in questions]
    
    answers = [div.findAll("span", {"class": "response-text"})[0].get_text().strip() if div.findAll("span", {"class": "response-text"}) else "n/a" for div in divs]
    
    for i in range(2, 12):
        ws.cell(row=index+1, column=i, value=answers[i-2])
    

print(questions)

for i in range(2, 12):
    ws.cell(row=1, column=i, value=questions[i-2])

    
wb.save("SurveyData.xlsx")


